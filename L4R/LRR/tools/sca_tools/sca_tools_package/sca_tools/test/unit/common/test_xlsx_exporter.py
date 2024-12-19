""" Tests for qac/exporters/xlsx_exporter.py """

from unittest.mock import patch
from unittest import TestCase, mock
from openpyxl import Workbook
from swq.common.export.xlsx_sheet import XlsxSheet
from swq.common.export.xlsx_exporter import XlsxExporter


class TestXlsxExporter(TestCase):
    """Test XlsxExporter Class"""
    def setUp(self):
        self.title = "title"

        with patch('swq.common.export.xlsx_exporter.LOGGER') as logger, \
            patch('os.path.basename') as basename, \
            patch('swq.common.export.xlsx_exporter.Workbook') as workbook:

            sheets = ["sheet1", "sheet2", "sheet3"]
            self.xlsx_sheet_sample = XlsxSheet(Workbook(self.title),
                                               self.title)
            instance = workbook.return_value
            instance.sheetnames = sheets
            instance.get_sheet_by_name.return_value = 0
            instance.remove_sheet.return_value = 0
            instance.create_sheet.return_value = self.xlsx_sheet_sample
            instance.save.return_value = 0

            basename.return_value = 0

            self.filepath = "file"
            self.xlsx_exporter = XlsxExporter(self.filepath)

            workbook.assert_called()
            instance.get_sheet_by_name.assert_has_calls([
                mock.call("sheet1"),
                mock.call("sheet2"),
                mock.call("sheet3")
            ])
            self.workbook_instance = instance
            instance.remove_sheet.assert_has_calls(
                [mock.call(0), mock.call(0),
                 mock.call(0)])
            basename.assert_called_with(self.filepath)
            logger.info.assert_called()

    @patch("swq.common.export.xlsx_exporter.Workbook", create=True)
    def test_create_sheet(self, workbook):
        """Test create_sheet() method"""
        instance = workbook.return_value

        instance.save.return_value = 0
        instance.create_sheet.return_value = 1
        workbook.create_sheet.return_value = 1

        with patch("swq.common.export.xlsx_exporter.XlsxSheet") as xlsx_sheet:
            xlsx_sheet.return_value = self.xlsx_sheet_sample
            return_value = self.xlsx_exporter.create_sheet(self.title)
            xlsx_sheet.assert_called_with(self.xlsx_sheet_sample, self.title)
            self.assertEqual(return_value, self.xlsx_sheet_sample)

    def test_save(self):
        """Test save() method"""
        return_value = self.xlsx_exporter.save()

        self.workbook_instance.save.assert_called_with(filename=self.filepath)
        self.assertEqual(return_value, self.filepath)
