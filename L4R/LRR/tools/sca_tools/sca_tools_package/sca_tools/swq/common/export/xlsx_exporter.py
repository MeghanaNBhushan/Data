# =============================================================================
#  C O P Y R I G H T
# -----------------------------------------------------------------------------
#  Copyright (c) 2022 by Robert Bosch GmbH. All rights reserved.
#
#  This file is property of Robert Bosch GmbH. Any unauthorized copy, use or
#  distribution is an offensive act against international law and may be
#  prosecuted under federal law. Its content is company confidential.
# =============================================================================
#  Filename: 	xslx_exporter.py
# ----------------------------------------------------------------------------
"""Defines a xlsx exporter implementation"""

from os import path
from openpyxl import Workbook
from openpyxl.styles import Alignment

from swq.common.logger import LOGGER
from swq.common.export.xlsx_sheet import XlsxSheet
from swq.common.export.spreadsheet_exporter import SpreadsheetExporter


class XlsxExporter(SpreadsheetExporter):
    """Defines a xlsx exporter"""

    __sheets = []

    def __init__(self, filepath):
        super().__init__(path.dirname(filepath), path.basename(filepath))
        self.__workbook = Workbook()
        # Ensures that the workbook is cleansed in all implementations
        # of the openpyxl library. Avoids problems later on create_sheet
        for sheet_name in self.__workbook.sheetnames:
            sheet = self.__workbook.get_sheet_by_name(sheet_name)
            self.__workbook.remove_sheet(sheet)
        self.__filepath = filepath

        LOGGER.info('Creating xlsx file in path %s', self.__filepath)

    def create_sheet(self, title: str):
        """Create a xlsx sheet"""
        result = XlsxSheet(self.__workbook.create_sheet(title), title)
        self.__sheets.append(result)

        return result

    def save(self):
        """Saves the xlsx state"""
        for sheet in self.__sheets:
            sheet.save()
        self.__workbook.save(filename=self.__filepath)
        return self.__filepath

    def wrap_text_in_cells(self):
        """Aligns multiline text in cells"""
        worksheet = self.__workbook.active
        for row in worksheet:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        self.__workbook.save(filename=self.__filepath)
        return self.__filepath
