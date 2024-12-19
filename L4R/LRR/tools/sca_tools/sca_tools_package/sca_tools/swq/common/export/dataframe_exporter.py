# =============================================================================
#  C O P Y R I G H T
# -----------------------------------------------------------------------------
#  Copyright (c) 2022 by Robert Bosch GmbH. All rights reserved.
#
#  This file is property of Robert Bosch GmbH. Any unauthorized copy, use or
#  distribution is an offensive act against international law and may be
#  prosecuted under federal law. Its content is company confidential.
# =============================================================================
#  Filename: 	dataframe_exporter.py
# ----------------------------------------------------------------------------
"""Defines a dataframe exporter implementation"""

from os import remove
from os.path import exists
from pandas import ExcelWriter
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

from swq.common.logger import LOGGER


def _set_column_width(work_sheet, column_names, col_widths):
    dim_holder = DimensionHolder(worksheet=work_sheet)
    for col in range(work_sheet.min_column, work_sheet.max_column + 1):
        col_name = column_names[col - 1]
        LOGGER.info('Column name found %s', col_name)
        col_width = col_widths[col_name]
        dim_holder[get_column_letter(col)] = ColumnDimension(work_sheet,
                                                             min=col,
                                                             max=col,
                                                             width=col_width)
    work_sheet.column_dimensions = dim_holder
    LOGGER.info('Column dimensions changed in %s', work_sheet.title)


class DataframeExporter():
    """Defines a dataframe to xlsx exporter"""
    __filepath = ''

    def __init__(self, filepath):
        self.__filepath = filepath
        LOGGER.info('Data frame exporter initialized with %s', filepath)
        self.__workbook = Workbook()
        if exists(self.__filepath):
            remove(self.__filepath)

    @property
    def workbook(self):
        """ Gets the workbook """
        return self.__workbook

    def _load_workbook(self, filepath):
        LOGGER.debug('Load workbook from %s', filepath)
        return load_workbook(filepath)

    def append_dataframe(self, sheet_name, df_data, cols=None):
        """ Appends a dataframe to a given work_sheet """
        LOGGER.info('Appending dataframe to work_sheet %s', sheet_name)
        mode = 'a+'
        if exists(self.__filepath):
            mode = 'a'

        # pylint: disable=abstract-class-instantiated:
        # this is a known and reported pylint issue,
        # see here https://github.com/PyCQA/pylint/issues/3060
        with ExcelWriter(self.__filepath, engine='openpyxl',
                         mode=mode) as writer:
            if cols is None:
                df_data.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                df_data.to_excel(writer,
                                 sheet_name=sheet_name,
                                 index=False,
                                 columns=cols)

        self.__workbook = self._load_workbook(self.__filepath)
        work_sheet = self.__workbook[sheet_name]
        LOGGER.info('Worksheet %s written to workbook %s', work_sheet.title,
                    self.__filepath)
        return work_sheet

    def format_columns(self, sheet_name, col_widths):
        """ Sets given column widths, saves immediatly """

        self.__workbook = self._load_workbook(self.__filepath)
        work_sheet = self.__workbook[sheet_name]
        LOGGER.info('Worksheet found %s', work_sheet.title)

        column_names = []
        for cell in work_sheet[1]:
            column_names.append(cell.value)

        LOGGER.info('column_names: %s', column_names)

        if len(column_names) == 0:
            LOGGER.error('Column names not found in worksheet %s', sheet_name)
        else:
            _set_column_width(work_sheet, column_names, col_widths)

        # add filters to columns
        work_sheet.auto_filter.ref = work_sheet.dimensions
        LOGGER.info('Auto filter set in %s', work_sheet.title)

        self.__workbook.save(self.__filepath)
        LOGGER.info('Workbook saved after change in %s', work_sheet.title)

        return work_sheet
