# -*- coding: utf-8 -*-
"""
Created on Tue Mar 23 18:11:21 2021

@author: YAH3KOR
"""
import argparse
from openpyxl import load_workbook
import sys
import csv

parser = argparse.ArgumentParser(description="Script to stash/Unstash jenkins artifacts")
parser.add_argument("-v", "--variant",
                        help='', nargs='?', const='FR5CU_DNNN1_NNN_N_XX_2_uC1')
args = parser.parse_args()
variant = str(args.variant)


from pathlib import Path
path1 = Path('../../generatedFiles/Radar_'+variant)

import logging
logger = logging.getLogger("")
logger.setLevel(logging.DEBUG)
    
logger_console_handler = logging.StreamHandler()
logger_console_handler.setLevel(logging.DEBUG)
    
log_format = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger_console_handler.setFormatter(log_format)

logger.addHandler(logger_console_handler)  

report = ''

logger.info("=======-Finding Calcres report ============")
for file in path1.glob('**/Calcres*.xlsx'):# ** in a path means any number of sub-directories in the path to match regex

    logger.info("Found : " + file.name)
    logger.info("Path  : " + str(file))
    report = load_workbook(filename=file)

logger.info("======= Searching for new libraries ============")
ram_rom = report['RAM_ROM_REPORT']
new_libs = []
default_composite = False

for row in ram_rom.iter_rows(values_only=True):
    if 'default_composite' in row :
        default_composite = True
        continue
    if default_composite :
        if row[0] != None :
            break
        else :
            if row[1] == None : 
                continue
            else :
                new_libs.append(row[1])
                logger.error("Unmapped library found : " + row[1])

if not default_composite : 
    logger.info("SUCCESS : No new libraries found !")     

logger.info("======= Searching for budget overshoots ============")           
summary = report['SUMMARY_REPORT']
budget_overshot = False
for row in summary.iter_rows(values_only=True):

        for value in list(row) :
            try : 
                if value != None and int(value) < 0 : 
                    budget_overshot = True
            except (ValueError) : 
                continue
    
if not budget_overshot : 
    logger.info("SUCCESS : All libraries within specified budget")
else : 
    logger.error("Budget overshoot found")
    logger.error("Budget Summary")
    for row in summary.iter_rows(values_only=True):
        logger.error(row)


# runs the csv_from_excel function:
if budget_overshot or default_composite : 
    sys.exit(-1)
else:
    ram_rom_detailed_csv = open('../../generatedFiles/Radar_'+variant+'/Calcres_Report_detailed_'+variant+'.csv', 'w')
    ram_rom_detailed_handler = csv.writer(ram_rom_detailed_csv , quoting=csv.QUOTE_ALL)

    for row in ram_rom.iter_rows(values_only=True):
        ram_rom_detailed_handler.writerow(row)

    ram_rom_detailed_csv.close()

    summary_csv = open('../../generatedFiles/Radar_'+variant+'/Calcres_Report_summary_'+variant+'.csv', 'w')
    summary_handler = csv.writer(summary_csv , quoting=csv.QUOTE_ALL)

    for row in summary.iter_rows(values_only=True):
        summary_handler.writerow(row)

    summary_csv.close()

    sys.exit(0)