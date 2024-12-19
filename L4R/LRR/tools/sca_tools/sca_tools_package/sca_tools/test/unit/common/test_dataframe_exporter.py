""" Tests for qac/exporters/xlsx_exporter.py """

from unittest.mock import patch
from unittest import TestCase, mock
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame, read_excel
from pandas.testing import assert_frame_equal
from swq.common.export.dataframe_exporter import DataframeExporter

GIVEN_COLUMNS = ('col1', 'col2')

_INPUT = {"col1": ("c1v1", "c1v2"), "col2": ("c2v1", "c2v2")}

COL_WIDTHS_SUMMARY = {"col1": 10, "col2": 60}

SHEET1 = "sheet1"

FILE_PATH = "file.xlsx"


def populate(sheet, df_data):
    rows = dataframe_to_rows(df_data, index=False, header=True)
    len_rows = 0

    for r_idx, row in enumerate(rows, 1):
        len_rows += 1
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx).value = value
    sheet.min_column = 0
    sheet.max_column = len(GIVEN_COLUMNS)


class TestDataframeExporter(TestCase):
    """Test DataframeExporter Class"""
    @patch('swq.common.export.dataframe_exporter.Workbook')
    @patch('swq.common.export.dataframe_exporter.LOGGER')
    def setUp(self, logger_mock, workbook_mock):
        self.expected_df_data = DataFrame(data=_INPUT, columns=GIVEN_COLUMNS)
        self.dataframe_exporter = DataframeExporter(FILE_PATH)

        workbook_mock.return_value.sheetnames = [SHEET1]
        work_sheet = workbook_mock.return_value[SHEET1]
        work_sheet.title = SHEET1
        work_sheet.values = _INPUT
        populate(work_sheet, self.expected_df_data)
        self.workbook_mock = workbook_mock

        self.workbook_mock.assert_called()
        logger_mock.info.assert_called()

    @patch('swq.common.export.dataframe_exporter.ExcelWriter')
    def test_append_dataframe(self, writer_mock):
        """Test append_dataframe() method"""

        writer_mock.engine = 'openpyxl'
        writer_mock.filepath = FILE_PATH
        writer_mock.mode = '+a'

        with patch.object(self.expected_df_data,
                               "to_excel") as to_excel_mock,\
            patch.object(self.dataframe_exporter, "_load_workbook")\
                 as load_workbook_mock:

            to_excel_mock.writer = writer_mock
            to_excel_mock.shee_name = SHEET1
            to_excel_mock.columns = GIVEN_COLUMNS

            load_workbook_mock.return_value = self.workbook_mock.return_value

            actual_worksheet = self.dataframe_exporter.append_dataframe(
                SHEET1, self.expected_df_data)

            to_excel_mock.assert_called_with(writer_mock().__enter__(),
                                             sheet_name=SHEET1,
                                             index=False)
            load_workbook_mock.assert_called_with(FILE_PATH)

            self.assertEqual(actual_worksheet.title, SHEET1)

            data = actual_worksheet.values
            actual_df_data = DataFrame(data, columns=GIVEN_COLUMNS)

            assert_frame_equal(actual_df_data.reset_index(drop=True),
                               self.expected_df_data.reset_index(drop=True))

    def test_format_columns(self):
        """Test format_columns() method"""
        with patch.object(self.dataframe_exporter, "_load_workbook")\
                 as load_workbook_mock,\
            patch.object(self.dataframe_exporter.workbook, "save")\
                as save_workbook_mock:
            load_workbook_mock.return_value = self.workbook_mock.return_value

            actual_worksheet = self.dataframe_exporter.format_columns(
                SHEET1, COL_WIDTHS_SUMMARY)

            self.assertEqual(actual_worksheet.title, SHEET1)
            load_workbook_mock.assert_called_with(FILE_PATH)
            save_workbook_mock.assert_called_with(FILE_PATH)
